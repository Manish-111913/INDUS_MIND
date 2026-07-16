"""Tabular parsing for the import engine (docs/05 S6).

Reads CSV or XLSX bytes into (headers, rows). XLSX is detected by its ZIP magic
(`PK`) and streamed read-only via openpyxl so a large sheet doesn't balloon memory.
"""

from __future__ import annotations

import csv
import io


def parse_table(data: bytes) -> tuple[list[str], list[list[str]]]:
    if data[:2] == b"PK":  # xlsx (zip container)
        return _parse_xlsx(data)
    return _parse_csv(data)


def _parse_csv(data: bytes) -> tuple[list[str], list[list[str]]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [list(r) for r in reader]
    if not rows:
        return [], []
    headers = [(h or "").strip() for h in rows[0]]
    body = [[(c if c is not None else "") for c in r] for r in rows[1:] if any(x for x in r)]
    return headers, body


def _parse_xlsx(data: bytes) -> tuple[list[str], list[list[str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    headers: list[str] = []
    body: list[list[str]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cells = ["" if c is None else str(c).strip() for c in row]
        if i == 0:
            headers = cells
        elif any(cells):
            body.append(cells)
    wb.close()
    return headers, body


def guess_mapping(headers: list[str], fields: list[str]) -> dict[str, str]:
    """Best-effort field → header map by normalised-name similarity."""
    def norm(s: str) -> str:
        return "".join(ch for ch in s.lower() if ch.isalnum())

    norm_headers = {norm(h): h for h in headers}
    mapping: dict[str, str] = {}
    for field in fields:
        nf = norm(field)
        if nf in norm_headers:
            mapping[field] = norm_headers[nf]
            continue
        # substring / prefix match either direction
        hit = next((h for nh, h in norm_headers.items() if nh and (nh in nf or nf in nh)), None)
        if hit is not None:
            mapping[field] = hit
    return mapping
